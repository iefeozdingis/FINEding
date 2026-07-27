import csv
import hashlib
import hmac
import logging
import secrets
import shutil
import sqlite3
import sys
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ui.money bilinçli olarak GUI-bağımsız saf bir modüldür (tkinter import
# etmez), bu yüzden veri katmanından güvenle kullanılabilir. Para ayrıştırma
# tek bir kaynakta tutulur; ikinci bir kopya tutmak sessiz para bozulmasına
# yol açıyordu (bkz. _tutar_parse).
from ui.money import para_parse

# ==========================
# Veritabanı Ayarları
# ==========================


def _uygulama_kok() -> Path:
    """Kalıcı veri dizininin kökünü döner.

    PyInstaller onefile paketinde __file__, çalıştırma anında açılan geçici
    _MEIxxxx klasörünü gösterir. O klasör her çalıştırmada YENİDEN oluşur ve
    çıkışta silinir; bu yüzden veritabanı oraya yazıldığında kullanıcının
    girdiği HER ŞEY uygulama kapanınca kayboluyor, her açılışta boş bir
    veritabanı geliyordu. Paketlenmiş uygulamada yol exe'nin yanına
    (sys.executable) sabitlenir.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    # Kaynaktan çalışırken: modül dizini. CWD'ye göreli olmamalı — kısayoldan
    # farklı dizinde başlatınca boş yeni bir veritabanı oluşturuyordu.
    return Path(__file__).resolve().parent


DB_FOLDER = _uygulama_kok() / "database"
DB_FOLDER.mkdir(exist_ok=True)

DB_PATH = DB_FOLDER / "finans.db"


class HmacAnahtarHatasi(Exception):
    """Yedek imzalama anahtarı okunamadığında/oluşturulamadığında fırlatılır."""


def _hmac_anahtari() -> bytes:
    """Kuruluma özel HMAC anahtarını okur; yoksa rastgele üretip saklar.

    Anahtar, veritabanı dosyasının yanında (yedeklerle BİRLİKTE taşınmayan
    database/ dizininde) tutulur; böylece sadece yedek dosyasına erişen biri
    checksum'ı yeniden üretip kurcalayamaz. Yol DB_PATH'ten türetilir ki
    testler farklı bir dizine izole olabilsin.
    """
    key_path = Path(DB_PATH).parent / ".hmac_key"
    try:
        if key_path.exists():
            return key_path.read_bytes()
        anahtar = secrets.token_bytes(32)
        key_path.write_bytes(anahtar)
        return anahtar
    except OSError as e:
        # Sabit gömülü anahtara DÜŞÜLMEZ: kaynak kodda herkese açık bir
        # anahtarla imzalamak, yedeği değiştiren birinin geçerli bir .hmac
        # üretip bütünlük kontrolünü geçmesini sağlıyordu — yani kurcalama
        # tespitini tamamen etkisiz kılıyordu. Anahtar üretilemiyorsa
        # imzalama/doğrulama yapılmaz.
        raise HmacAnahtarHatasi(
            f"Yedek imzalama anahtarı okunamadı/oluşturulamadı: {e}"
        ) from e


# Güvenli şifre hash'leme

try:
    import bcrypt
    _HAS_BCRYPT = True
except ImportError:
    _HAS_BCRYPT = False


def _sifre_hashla(sifre: str) -> str:
    """Şifreyi bcrypt ile hash'ler.

    bcrypt yoksa ARTIK SESSİZCE ZAYIF HASH ÜRETİLMEZ. Eski yedek yol
    tuzsuz SHA-256 + kaynak koda gömülü sabit "pepper" kullanıyordu: aynı
    şifre her kullanıcıda aynı hash'i veriyor, DB'yi okuyan biri gökkuşağı
    tablosu/GPU ile saniyeler içinde kırabiliyordu. bcrypt zorunlu bir
    bağımlılık (requirements.txt), yokluğu bir kurulum hatasıdır.
    """
    if not _HAS_BCRYPT:
        raise RuntimeError(
            "bcrypt kurulu değil; güvenli şifre saklanamıyor. "
            "Kurulum için: pip install -r requirements.txt"
        )
    return bcrypt.hashpw(sifre.encode(), bcrypt.gensalt()).decode()


def _sifre_dogrula(sifre: str, hash_deger: str) -> bool:
    """Şifre ile hash'i karşılaştırır (bcrypt veya eski SHA-256 hash'ler için)."""
    if hash_deger.startswith("$2"):
        if _HAS_BCRYPT:
            return bcrypt.checkpw(sifre.encode(), hash_deger.encode())
        return False
    # Eski (bcrypt öncesi) SHA-256 hash'ler YALNIZCA DOĞRULAMA için hâlâ
    # destekleniyor: mevcut kullanıcılar giriş yapabilsin ve hash'leri
    # kullanici_dogrula içindeki upgrade-on-login ile bcrypt'e yükselsin.
    # Yeni hash üretimi için bu yol artık KULLANILMIYOR (_sifre_hashla).
    legacy_hash = hashlib.sha256(b"Fineding2024!" + sifre.encode()).hexdigest()
    return hmac.compare_digest(legacy_hash, hash_deger)


def csv_guvenli(deger: Any) -> Any:
    """CSV/Excel formül enjeksiyonuna karşı hücreyi temizler.

    =, +, -, @, TAB veya CR ile başlayan metinler Excel/LibreOffice'te
    formül (=HYPERLINK, DDE vb.) olarak çalışır. Bu tür hücrelerin önüne
    tek tırnak eklenerek düz metin olması sağlanır.
    """
    if isinstance(deger, str) and deger and deger[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + deger
    return deger


def para_kurus(lira: Any) -> int:
    """Lira-float değerini tam sayı KURUŞ'a çevirir — YAZMA sınırı.

    Para artık DB'de kuruş bazlı INTEGER saklanır. Önceden REAL (float)
    saklandığı için 0.1+0.2 sınıfı birikimli yuvarlama hataları ve çok
    satırlı SUM kayması oluşabiliyordu; kuruş tam sayısı bu sınıfı kökten
    bitirir ve çoklu para birimi (minör birim = tam sayı) için zemin hazırlar.

    Kullanıcı girdisi (para_parse) en fazla 2 ondalık ürettiği için
    tutar*100 matematiksel olarak tam sayıdır; round() yalnızca float
    gösterim gürültüsünü (ör. 19.99*100=1998.9999…) temizler.
    """
    return int(round(float(lira) * 100))


def para_lira(kurus: Any) -> float:
    """Tam sayı kuruşu lira-float'a çevirir — OKUMA sınırı.

    Public API (UI'ye bakan) lira-float döner; depolama kuruş olsa da
    arayüz sözleşmesi değişmez. Saklanan değer tam olduğundan (tam sayı
    kuruş) yalnızca gösterim için buradan float'a dönülür.
    """
    return (kurus or 0) / 100


# islemler tablosu için okuma projeksiyonu. tutar KURUŞ (INTEGER) saklanır;
# okuma sınırında lira'ya (÷100.0) çevrilir. Kolon SIRASI create_tables ile
# birebir aynıdır (UI indeksleri islem[5]=tutar'a dayanır). Sabit literaldir,
# enjeksiyon riski yoktur.
#
# Sona eklenen çoklu-para kolonları (index 8-9): para_birimi ve orijinal tutar
# (lira). COALESCE(orijinal_tutar, tutar): eski/iç-yol kayıtları orijinal_tutar
# vermese bile TL tutarına düşer, yani TL işlemler doğru görünür.
_ISLEM_SECIM = (
    "id, tarih, tur, kategori, aciklama, tutar/100.0, etiketler, "
    "kullanici_id, para_birimi, COALESCE(orijinal_tutar, tutar)/100.0"
)


def normalize_date(tarih_str: str) -> str:
    """Normalize a date string to ISO YYYY-MM-DD.

    Accepts DD.MM.YYYY or YYYY-MM-DD and returns YYYY-MM-DD.
    Raises ValueError on invalid formats.
    """
    if not isinstance(tarih_str, str):
        raise ValueError("Tarih string olmalidir")

    tarih_str = tarih_str.strip()
    # DD.MM.YYYY
    try:
        if "." in tarih_str:
            dt = datetime.strptime(tarih_str, "%d.%m.%Y")
            return dt.strftime("%Y-%m-%d")
        # YYYY-MM-DD
        dt = datetime.strptime(tarih_str, "%Y-%m-%d")
        return dt.strftime("%Y-%m-%d")
    except Exception as e:
        raise ValueError(f"Geçersiz tarih formatı: {tarih_str}") from e


# Şema sürümü — her artışta _migrate() ilgili adımı uygular
SCHEMA_VERSION = 8

# Geri-al yığınında en fazla kaç silme partisi tutulur
GERI_AL_YIGIN_SINIRI = 20

# Minimum şifre uzunluğu (veri katmanında zorlanır)
MIN_SIFRE_UZUNLUK = 8


class YetkiHatasi(Exception):
    """Yetkisiz bir yönetim işlemi denendiğinde fırlatılır."""


# Kullanıcıya ait finans tabloları (izolasyon için kullanici_id taşır)
_KULLANICI_TABLOLARI = (
    "islemler",
    "butceler",
    "borclar",
    "planlanan",
    "tasarruf_hedefleri",
    "tekrarlayan",
)


from db.borc import BorcMixin  # noqa: E402
from db.kategori import KategoriMixin  # noqa: E402
from db.kullanici import KullaniciMixin  # noqa: E402
from db.planlama import PlanlamaMixin  # noqa: E402
from db.tasarruf import TasarrufMixin  # noqa: E402
from db.tekrarlayan import TekrarlayanMixin  # noqa: E402


class Database(
    PlanlamaMixin,
    BorcMixin,
    KullaniciMixin,
    KategoriMixin,
    TekrarlayanMixin,
    TasarrufMixin,
):
    def __init__(self, kullanici_id: int = 1) -> None:
        self.conn = self._baglan()
        self.cursor = self.conn.cursor()
        # Geri-al geçmişi: her eleman bir SİLME PARTİSİ (bir veya çok kayıt).
        # Tek slot yerine yığın: toplu silmede yalnızca son kayıt geri
        # alınabiliyor, kalanlar sessizce kalıcı kayboluyordu.
        self._silinen_yigin: List[List[Any]] = []
        # v5 migrasyonunda bakiyeden ayrılan eski borç işlemi sayısı (bilgi)
        self.migrate_silinen_borc_islem = 0
        # Aktif oturum kullanıcısı — tüm finans sorguları bununla filtrelenir.
        # Varsayılan 1 (ilk/admin kullanıcı); giriş sonrası oturum_ac ile
        # gerçek kullanıcıya ayarlanır.
        self.aktif_kullanici_id = kullanici_id
        self.create_tables()
        self._migrate()
        self._index_olustur()

    def oturum_ac(self, kullanici_id: int) -> None:
        """Aktif oturum kullanıcısını ayarlar; sonraki tüm sorgular bu
        kullanıcının verisiyle sınırlanır."""
        self.aktif_kullanici_id = int(kullanici_id)

    @staticmethod
    def _baglan() -> sqlite3.Connection:
        """Ortak bağlantı ayarlarıyla SQLite bağlantısı açar.

        busy_timeout: arka plan thread'leri (tekrarlayan/borç kontrolü) ve UI
        aynı anda yazınca 'database is locked' hatası yerine kısa süre bekler.
        """
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=10000")
        conn.execute("PRAGMA foreign_keys=ON")
        return conn

    @contextmanager
    def _transaction(self):
        """Çok adımlı yazmaları atomik yapar.

        Korumasız hali (yalnızca sonda commit) hata anında bekleyen kısmi
        değişiklikleri geri almıyordu; açık kalan transaction ilgisiz bir
        sonraki commit() ile sessizce kalıcılaşabiliyordu.
        """
        self.cursor.execute("BEGIN")
        try:
            yield self.cursor
        except Exception:
            self.conn.rollback()
            raise
        else:
            self.conn.commit()

    def _migrate(self) -> None:
        """PRAGMA user_version tabanlı numaralı şema migrasyonu.

        'dene-yut' ALTER TABLE yerine sürüm numarasıyla ilerleyen, gerçek
        hataları yutmayan bir yol. Mevcut normalize edilmemiş borç
        tarihlerini de ISO'ya çevirir.
        """
        mevcut = self.conn.execute("PRAGMA user_version").fetchone()[0]
        if mevcut < 1:
            self._migrate_borc_tarihleri()
        if mevcut < 2:
            self._migrate_kullanici_id()
        if mevcut < 3:
            self._migrate_giris_denemeleri()
        if mevcut < 4:
            self._migrate_kategori_izolasyonu()
        if mevcut < 5:
            self._migrate_borc_bakiyeden_ayir()
        if mevcut < 6:
            self._migrate_borc_odemeler_fk()
        if mevcut < 7:
            self._migrate_kurus_tam_sayi()
        if mevcut < 8:
            self._migrate_para_birimi()
        if mevcut < SCHEMA_VERSION:
            self.conn.execute(f"PRAGMA user_version={SCHEMA_VERSION}")
            self.conn.commit()

    def _migrate_kullanici_id(self) -> None:
        """Finans tablolarına kullanici_id kolonu ekler; mevcut tüm veriyi
        ilk kullanıcıya (admin, id=1) atar. Böylece çok kullanıcılı giriş
        artık gerçek veri izolasyonu sağlar (önceden tüm kullanıcılar aynı
        havuzu görüyordu)."""
        for tablo in _KULLANICI_TABLOLARI:
            if tablo == "butceler":
                continue  # aşağıda özel olarak yeniden yapılandırılır
            kolonlar = {
                r[1]
                for r in self.conn.execute(f"PRAGMA table_info({tablo})").fetchall()
            }
            if "kullanici_id" not in kolonlar:
                self.conn.execute(
                    f"ALTER TABLE {tablo} ADD COLUMN kullanici_id INTEGER DEFAULT 1"
                )
                self.conn.execute(
                    f"UPDATE {tablo} SET kullanici_id=1 WHERE kullanici_id IS NULL"
                )
        self._migrate_butce_kullanici()
        self.conn.commit()

    def _migrate_giris_denemeleri(self) -> None:
        """v3: eksik kolonları eski veritabanlarına ekler.

        - kullanicilar.basarisiz_deneme / son_basarisiz: kaba kuvvet sayacı
          artık kalıcı. Önceden yalnızca giriş penceresinin bellek alanındaydı;
          pencereyi kapatıp açmak gecikmeyi sıfırlıyordu.
        - etiketler / aktarim_tarihi / son_islenen_donem: bu kolonlar önceden
          create_tables içinde 'dene-yut' ALTER ile ekleniyordu. Şema artık
          create_tables'ta eksiksiz tanımlı; eski DB'ler buradan yükseltilir.
        """
        eklenecek = [
            ("kullanicilar", "basarisiz_deneme", "INTEGER DEFAULT 0"),
            ("kullanicilar", "son_basarisiz", "TEXT"),
            ("islemler", "etiketler", "TEXT DEFAULT ''"),
            ("planlanan", "aktarim_tarihi", "TEXT DEFAULT ''"),
            ("tekrarlayan", "son_islenen_donem", "TEXT DEFAULT ''"),
        ]
        for tablo, kolon, tip in eklenecek:
            mevcut = {
                r[1]
                for r in self.conn.execute(
                    f"PRAGMA table_info({tablo})"
                ).fetchall()
            }
            if not mevcut:
                continue  # tablo yok (yeni kurulum) — create_tables halleder
            if kolon not in mevcut:
                self.conn.execute(
                    f"ALTER TABLE {tablo} ADD COLUMN {kolon} {tip}"
                )
        self.conn.commit()

    def _migrate_borc_odemeler_fk(self) -> None:
        """v6: borc_odemeler'e borc_id → borclar(id) FK'sini ekler.

        SQLite'ta var olan bir tabloya ALTER ile FK eklenemez; tablo FK'li
        olarak yeniden kurulup veri kopyalanır. Yetim ödeme satırları (borcu
        artık olmayan) kopyalanmadan atılır — zaten geçersizler.
        SQLite dokümanının önerdiği gibi rebuild sırasında foreign_keys
        geçici kapatılır.
        """
        # Tablo yoksa (yeni kurulum create_tables halleder) atla
        var = self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='borc_odemeler'"
        ).fetchone()
        if not var:
            return
        # Zaten FK varsa tekrar kurma
        fkler = self.conn.execute(
            "PRAGMA foreign_key_list(borc_odemeler)"
        ).fetchall()
        if fkler:
            return

        self.conn.execute("PRAGMA foreign_keys=OFF")
        try:
            self.conn.execute("BEGIN")
            self.conn.execute("""
                CREATE TABLE borc_odemeler_yeni(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    borc_id INTEGER NOT NULL,
                    tarih TEXT NOT NULL,
                    tutar REAL NOT NULL,
                    FOREIGN KEY (borc_id) REFERENCES borclar(id) ON DELETE CASCADE
                )
            """)
            # Yalnızca hâlâ var olan borçlara ait ödemeleri taşı (yetimleri at)
            self.conn.execute("""
                INSERT INTO borc_odemeler_yeni (id, borc_id, tarih, tutar)
                SELECT id, borc_id, tarih, tutar FROM borc_odemeler
                WHERE borc_id IN (SELECT id FROM borclar)
            """)
            self.conn.execute("DROP TABLE borc_odemeler")
            self.conn.execute(
                "ALTER TABLE borc_odemeler_yeni RENAME TO borc_odemeler"
            )
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise
        finally:
            self.conn.execute("PRAGMA foreign_keys=ON")

    def _migrate_kurus_tam_sayi(self) -> None:
        """v7: Para REAL(lira) → INTEGER(kuruş). Her para sütununu 100 ile
        çarpıp tam sayıya yuvarlar (1 TL = 100 kuruş).

        Amaç: depolama ve SUM tam kesinlik kazansın; 0.1+0.2 sınıfı float
        birikim hataları bitsin; çoklu para birimi (minör birim = tam sayı)
        için zemin hazırlansın. Yeni kurulumlar zaten INTEGER şemayla açılır;
        bu adım yalnızca eski (REAL/lira) veritabanlarındaki değerleri çevirir.

        Not: Eski DB'lerde sütun AFİNİTESİ REAL kalır (SQLite'ta afinite
        rebuild olmadan değişmez) ama artık tam sayı kuruş tutar — bu
        değerler 2^53'e kadar float'ta da tam temsil edildiğinden depolama ve
        toplama kesin kalır. Tüm yazmalar para_kurus'tan geçtiği için bir daha
        kesirli değer yazılamaz. Çekirdek tabloları (islemler dahil) marjinal
        bir afinite kozmetiği için yeniden kurmak orantısız risk taşıdığından
        değerler yerinde çevrilir.
        """
        donusumler = [
            ("islemler", ("tutar",)),
            ("butceler", ("tutar",)),
            ("planlanan", ("tutar",)),
            ("borclar", ("toplam_tutar", "kalan_tutar")),
            ("tekrarlayan", ("tutar",)),
            ("borc_odemeler", ("tutar",)),
            ("tasarruf_hedefleri", ("hedef_tutar", "biriken_tutar")),
        ]
        for tablo, kolonlar in donusumler:
            var = self.conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (tablo,),
            ).fetchone()
            if not var:
                continue  # tablo yok (kısmi/eski kurulum) — create_tables halleder
            for kolon in kolonlar:
                # CAST(ROUND(x*100)) — 19.99 gibi kayan değerlerdeki gösterim
                # gürültüsünü (1998.9999…) yuvarlayıp tam kuruşa oturtur.
                self.conn.execute(
                    f"UPDATE {tablo} SET {kolon} = "
                    f"CAST(ROUND({kolon} * 100) AS INTEGER) "
                    f"WHERE {kolon} IS NOT NULL"
                )
        self.conn.commit()

    def _migrate_para_birimi(self) -> None:
        """v8: islemler'e çoklu para birimi kolonları ekler.

        - para_birimi: işlemin orijinal para birimi (varsayılan 'TRY').
        - orijinal_tutar: orijinal birimdeki tutar, KURUŞ/cent (minör birim).

        tutar sütunu TL-kuruş (temel/raporlama) olarak KALIR; yabancı işlemler
        girişte TCMB kuruyla TL'ye çevrilip öyle saklanır. Böylece tüm mevcut
        SUM/bakiye/bütçe/rapor sorguları değişmeden TL bazında doğru çalışır.
        Eski satırların hepsi TL kabul edilir: orijinal_tutar = tutar.

        v7'den (kuruş) sonra çalışır; orijinal_tutar da kuruş olur ve tutar ile
        birebir eşleşir.
        """
        kolonlar = {
            r[1]
            for r in self.conn.execute(
                "PRAGMA table_info(islemler)"
            ).fetchall()
        }
        if not kolonlar:
            return  # tablo yok (yeni kurulum) — create_tables halleder
        if "para_birimi" not in kolonlar:
            self.conn.execute(
                "ALTER TABLE islemler ADD COLUMN para_birimi TEXT NOT NULL "
                "DEFAULT 'TRY'"
            )
        if "orijinal_tutar" not in kolonlar:
            self.conn.execute(
                "ALTER TABLE islemler ADD COLUMN orijinal_tutar INTEGER"
            )
            self.conn.execute(
                "UPDATE islemler SET orijinal_tutar = tutar "
                "WHERE orijinal_tutar IS NULL"
            )
        self.conn.commit()

    def _migrate_borc_bakiyeden_ayir(self) -> None:
        """v5: borç/alacağı ana bakiyeden ayır (muhasebe modeli B).

        Eski model her borç/alacak ödemesini islemler'e 'borc-odeme' etiketli
        bir Gelir/Gider olarak yazıyordu; bu, bakiyeyi anapara kadar şişiriyor
        ve kredili alışverişte çift sayıma yol açıyordu. Bu satırları islemler'den
        siliyoruz — ÖDEME GEÇMİŞİ borc_odemeler tablosunda korunur, yani veri
        kaybı yok; yalnızca bakiyeyi bozan türev kayıtlar temizlenir.

        Kaç satır silindiği self.migrate_silinen_borc_islem'e yazılır (bilgi/log).
        """
        try:
            cur = self.conn.execute(
                "SELECT COUNT(*) FROM islemler WHERE etiketler='borc-odeme'"
            )
            adet = cur.fetchone()[0]
        except sqlite3.OperationalError:
            adet = 0  # islemler/etiketler yoksa (yeni kurulum) yapılacak iş yok
        if adet:
            self.conn.execute(
                "DELETE FROM islemler WHERE etiketler='borc-odeme'"
            )
            logging.getLogger(__name__).info(
                "Borç/alacak modeli B: %d 'borc-odeme' işlemi bakiyeden ayrıldı "
                "(geçmiş borc_odemeler'de korunuyor).", adet
            )
        self.migrate_silinen_borc_islem = adet
        self.conn.commit()

    def _migrate_kategori_izolasyonu(self) -> None:
        """v4: global özel-kategori ayarlarını admin'e (id=1) taşır.

        Kategori anahtarı 'kategoriler_gelir'/'kategoriler_gider' global idi;
        artık kullanıcıya özel ('..._<uid>'). Mevcut global kategoriler
        kaybolmasın diye ilk kullanıcının (admin) anahtarına aktarılır.
        """
        for tur in ("gelir", "gider"):
            eski = f"kategoriler_{tur}"
            yeni = f"kategoriler_{tur}_1"
            deger = self.conn.execute(
                "SELECT deger FROM ayarlar WHERE anahtar=?", (eski,)
            ).fetchone()
            if deger and deger[0]:
                # Admin'de zaten varsa üzerine yazma; yoksa taşı
                var = self.conn.execute(
                    "SELECT 1 FROM ayarlar WHERE anahtar=?", (yeni,)
                ).fetchone()
                if not var:
                    self.conn.execute(
                        "INSERT INTO ayarlar (anahtar, deger) VALUES (?, ?)",
                        (yeni, deger[0]),
                    )
                self.conn.execute("DELETE FROM ayarlar WHERE anahtar=?", (eski,))
        self.conn.commit()

    def _migrate_butce_kullanici(self) -> None:
        """butceler'in UNIQUE kısıtı (ay,yil,kategori) kullanıcı izolasyonunu
        engelliyordu; (ay,yil,kategori,kullanici_id) olacak şekilde tabloyu
        yeniden kurar ve mevcut satırları admin'e (id=1) atar."""
        kolonlar = {
            r[1]
            for r in self.conn.execute("PRAGMA table_info(butceler)").fetchall()
        }
        if "kullanici_id" in kolonlar:
            return  # zaten yeni şema
        self.conn.executescript(
            """
            ALTER TABLE butceler RENAME TO butceler_eski;
            CREATE TABLE butceler(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ay INTEGER NOT NULL,
                yil INTEGER NOT NULL,
                kategori TEXT NOT NULL,
                tutar REAL NOT NULL,
                kullanici_id INTEGER DEFAULT 1,
                UNIQUE(ay, yil, kategori, kullanici_id)
            );
            INSERT INTO butceler (id, ay, yil, kategori, tutar, kullanici_id)
                SELECT id, ay, yil, kategori, tutar, 1 FROM butceler_eski;
            DROP TABLE butceler_eski;
            """
        )

    def _migrate_borc_tarihleri(self) -> None:
        """Eski GG.AA.YYYY borç tarihlerini ISO YYYY-MM-DD'ye çevirir."""
        self.cursor.execute(
            "SELECT id, baslangic_tarih, vade_tarih FROM borclar"
        )
        for bid, bas, vade in self.cursor.fetchall():
            yeni_bas = self._iso_veya_ayni(bas)
            yeni_vade = self._iso_veya_ayni(vade)
            if yeni_bas != bas or yeni_vade != vade:
                self.conn.execute(
                    "UPDATE borclar SET baslangic_tarih=?, vade_tarih=? WHERE id=?",
                    (yeni_bas, yeni_vade, bid),
                )

    @staticmethod
    def _iso_veya_ayni(tarih: Any) -> Any:
        """GG.AA.YYYY ise ISO'ya çevirir, aksi halde olduğu gibi bırakır."""
        if not tarih or not isinstance(tarih, str) or "." not in tarih:
            return tarih
        try:
            return normalize_date(tarih)
        except ValueError:
            return tarih

    def _index_olustur(self) -> None:
        """Sık kullanılan filtre kolonlarına index ekler (yoksa)."""
        self.conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_islemler_tarih ON islemler(tarih)"
        )
        self.conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_islemler_tur_tarih "
            "ON islemler(tur, tarih)"
        )
        self.conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_islemler_kullanici "
            "ON islemler(kullanici_id)"
        )
        self.conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_islemler_kategori "
            "ON islemler(kategori)"
        )
        self.conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_borclar_vade "
            "ON borclar(durum, vade_tarih)"
        )
        self.conn.commit()

    # ==========================
    # Tablolar
    # ==========================

    def create_tables(self) -> None:
        """GÜNCEL şemayı kurar (yalnızca CREATE TABLE IF NOT EXISTS).

        Kolon eklemeleri buraya DEĞİL, numaralı migrasyona aittir. Önceden
        şema ikiye bölünmüştü: create_tables tabloları kullanici_id olmadan
        kuruyor, kolon yalnızca _migrate'ten geliyordu. Bu yüzden migrasyon
        atlanan/kesilen bir kurulumda tüm sorgular 'no such column' ile
        patlıyordu ve şemayı okuyan geliştirici eksik bir tablo görüyordu.
        """
        # PARA BİRİMİ: Tüm tutar sütunları INTEGER ve KURUŞ cinsindendir
        # (1 TL = 100). REAL(lira) float birikim/SUM hatalarına yol açıyordu;
        # kuruş tam sayısı depolama ve toplamayı tam kesin yapar. DB sınırında
        # para_kurus (yaz) / para_lira (oku) ile lira-float'a çevrilir; public
        # API ve UI değişmeden lira ile çalışmaya devam eder.
        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS islemler(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tarih TEXT NOT NULL,
            tur TEXT NOT NULL,
            kategori TEXT NOT NULL,
            aciklama TEXT,
            tutar INTEGER NOT NULL,
            etiketler TEXT DEFAULT '',
            kullanici_id INTEGER DEFAULT 1,
            para_birimi TEXT NOT NULL DEFAULT 'TRY',
            orijinal_tutar INTEGER
        )
        """)

        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS butceler(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ay INTEGER NOT NULL,
            yil INTEGER NOT NULL,
            kategori TEXT NOT NULL,
            tutar INTEGER NOT NULL,
            kullanici_id INTEGER DEFAULT 1,
            UNIQUE(ay, yil, kategori, kullanici_id)
        )
        """)

        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS ayarlar(
            anahtar TEXT PRIMARY KEY,
            deger TEXT
        )
        """)

        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS planlanan(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ay INTEGER NOT NULL,
            yil INTEGER NOT NULL,
            kategori TEXT NOT NULL,
            tur TEXT NOT NULL,
            aciklama TEXT,
            tutar INTEGER NOT NULL,
            aktarim_tarihi TEXT DEFAULT '',
            kullanici_id INTEGER DEFAULT 1
        )
        """)

        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS borclar(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tur TEXT NOT NULL,
            aciklama TEXT NOT NULL,
            kisi TEXT,
            toplam_tutar INTEGER NOT NULL,
            kalan_tutar INTEGER NOT NULL,
            baslangic_tarih TEXT,
            vade_tarih TEXT,
            durum TEXT NOT NULL DEFAULT 'Aktif',
            kullanici_id INTEGER DEFAULT 1
        )
        """)

        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS kullanicilar(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kullanici_adi TEXT UNIQUE NOT NULL,
            sifre_hash TEXT NOT NULL,
            ad_soyad TEXT,
            olusturma_tarihi TEXT NOT NULL,
            basarisiz_deneme INTEGER DEFAULT 0,
            son_basarisiz TEXT
        )
        """)

        # Tekrarlayan işlemler
        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS tekrarlayan(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tur TEXT NOT NULL,
            kategori TEXT NOT NULL,
            aciklama TEXT,
            tutar INTEGER NOT NULL,
            gun INTEGER NOT NULL,
            aktif INTEGER NOT NULL DEFAULT 1,
            son_islenen_donem TEXT DEFAULT '',
            kullanici_id INTEGER DEFAULT 1
        )
        """)

        # İşlem geçmişi (audit log)
        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS islem_log(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            zaman TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            islem_turu TEXT NOT NULL,
            islem_id INTEGER,
            detay TEXT
        )
        """)

        # Borç/alacak ödeme geçmişi.
        # FK: borç silinince ödemeleri de otomatik silinsin (yetim kayıt yok).
        # PRAGMA foreign_keys=ON (bkz. _baglan) ile birlikte DB seviyesinde
        # bütünlük sağlar; elle temizlemeye güvenmez.
        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS borc_odemeler(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            borc_id INTEGER NOT NULL,
            tarih TEXT NOT NULL,
            tutar INTEGER NOT NULL,
            FOREIGN KEY (borc_id) REFERENCES borclar(id) ON DELETE CASCADE
        )
        """)

        # Tasarruf hedefleri
        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS tasarruf_hedefleri(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ad TEXT NOT NULL,
            hedef_tutar INTEGER NOT NULL,
            biriken_tutar INTEGER NOT NULL DEFAULT 0,
            hedef_tarih TEXT,
            kullanici_id INTEGER DEFAULT 1
        )
        """)

        # İlk kullanıcı otomatik admin olur (ID=1)

        self.conn.commit()

    # ==========================
    # İŞLEM LOG KAYDI
    # ==========================

    def _log_islem(self, islem_turu: str, islem_id: Any = None, detay: str = "") -> None:
        """İşlemi audit log'a yazar (commit ETMEZ — çağıranın transaction'ına
        dahildir; böylece kayıt + log atomik olur)."""
        self.cursor.execute(
            "INSERT INTO islem_log (islem_turu, islem_id, detay) VALUES (?,?,?)",
            (islem_turu, islem_id, detay),
        )

    # ==========================
    # GELİR EKLE
    # ==========================

    def gelir_ekle(
        self, tarih: str, kategori: str, aciklama: Optional[str], tutar: float,
        etiketler: str = "", para_birimi: str = "TRY",
        orijinal_tutar: Optional[float] = None,
    ) -> None:
        """Gelir ekler. tutar TEMEL (TL) değerdir; yabancı para birimi için
        çağıran (UI) kur çevirisini yapıp tutar=orijinal*kur geçirir,
        orijinal_tutar'da da orijinal birimdeki tutarı (lira) verir."""
        tarih_iso = normalize_date(tarih)
        orijinal = tutar if orijinal_tutar is None else orijinal_tutar
        self.cursor.execute(
            """
        INSERT INTO islemler
        (tarih,tur,kategori,aciklama,tutar,etiketler,kullanici_id,
         para_birimi,orijinal_tutar)
        VALUES (?,?,?,?,?,?,?,?,?)
        """,
            (tarih_iso, "Gelir", kategori, aciklama, para_kurus(tutar),
             etiketler, self.aktif_kullanici_id, para_birimi,
             para_kurus(orijinal)),
        )
        self._log_islem("gelir_ekle", self.cursor.lastrowid, f"{kategori}: {tutar}")
        self.conn.commit()

    # ==========================
    # GİDER EKLE
    # ==========================

    def gider_ekle(
        self, tarih: str, kategori: str, aciklama: Optional[str], tutar: float,
        etiketler: str = "", para_birimi: str = "TRY",
        orijinal_tutar: Optional[float] = None,
    ) -> None:
        """Gider ekler. tutar TEMEL (TL) değerdir; yabancı para birimi için
        çağıran (UI) kur çevirisini yapıp tutar=orijinal*kur geçirir,
        orijinal_tutar'da da orijinal birimdeki tutarı (lira) verir."""
        tarih_iso = normalize_date(tarih)
        orijinal = tutar if orijinal_tutar is None else orijinal_tutar
        self.cursor.execute(
            """
        INSERT INTO islemler
        (tarih,tur,kategori,aciklama,tutar,etiketler,kullanici_id,
         para_birimi,orijinal_tutar)
        VALUES (?,?,?,?,?,?,?,?,?)
        """,
            (tarih_iso, "Gider", kategori, aciklama, para_kurus(tutar),
             etiketler, self.aktif_kullanici_id, para_birimi,
             para_kurus(orijinal)),
        )
        self._log_islem("gider_ekle", self.cursor.lastrowid, f"{kategori}: {tutar}")
        self.conn.commit()

    # ==========================
    # TÜM İŞLEMLER
    # ==========================

    def tum_islemler(self) -> List[Tuple[Any, ...]]:
        self.cursor.execute(
            f"SELECT {_ISLEM_SECIM} FROM islemler WHERE kullanici_id=? "
            "ORDER BY id DESC",
            (self.aktif_kullanici_id,),
        )
        return self.cursor.fetchall()

    def tum_islem_sayisi(self) -> int:
        self.cursor.execute(
            "SELECT COUNT(*) FROM islemler WHERE kullanici_id=?",
            (self.aktif_kullanici_id,),
        )
        row = self.cursor.fetchone()
        return row[0] if row else 0

    def islem_ara(
        self, arama: str = "", tur: str = "", limit: Optional[int] = None,
        donem: str = "",
    ) -> List[Tuple[Any, ...]]:
        """Metin, tür ve döneme göre işlemleri filtreleyerek arar.

        limit verilirse en yeni N kayıt döner (dashboard performansı için).
        donem: "" (tümü), "bugun" veya "hafta". Üç filtre BİRLİKTE uygulanır;
        önceden dönem filtresi ayrı bir sorgu yoluydu ve aramayı yok sayıyordu.
        """
        sorgu = f"SELECT {_ISLEM_SECIM} FROM islemler WHERE kullanici_id=?"
        params: List[Any] = [self.aktif_kullanici_id]
        if donem == "bugun":
            from datetime import date
            sorgu += " AND tarih=?"
            params.append(date.today().strftime("%Y-%m-%d"))
        elif donem == "hafta":
            from datetime import date, timedelta
            bugun = date.today()
            hafta_basi = bugun - timedelta(days=bugun.weekday())
            sorgu += " AND tarih BETWEEN ? AND ?"
            params.extend([
                hafta_basi.strftime("%Y-%m-%d"), bugun.strftime("%Y-%m-%d")
            ])
        if arama:
            sorgu += (
                " AND (kategori LIKE ? ESCAPE '\\' OR aciklama LIKE ? ESCAPE '\\'"
                " OR printf('%.2f', tutar/100.0) LIKE ? ESCAPE '\\'"
                " OR etiketler LIKE ? ESCAPE '\\')"
            )
            # Kullanıcının yazdığı % ve _ joker karakter olarak yorumlanmasın
            kacisli = arama.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            like = f"%{kacisli}%"
            params.extend([like, like, like, like])
        if tur:
            sorgu += " AND tur=?"
            params.append(tur)
        sorgu += " ORDER BY id DESC"
        if limit is not None:
            sorgu += " LIMIT ?"
            params.append(int(limit))
        self.cursor.execute(sorgu, tuple(params))
        return self.cursor.fetchall()

    def guncelle_islem(
        self,
        id: int,
        tarih: str,
        tur: str,
        kategori: str,
        aciklama: Optional[str],
        tutar: float,
        etiketler: Optional[str] = None,
        para_birimi: str = "TRY",
        orijinal_tutar: Optional[float] = None,
    ) -> None:
        """İşlemi günceller. tutar TEMEL (TL) değerdir; yabancı para için
        çağıran (UI) güncel kurla tutar=orijinal*kur geçirir, orijinal_tutar'da
        orijinal birimdeki tutarı verir. Para birimi belirtilmezse TL kabul
        edilir (orijinal_tutar=tutar)."""
        tarih_iso = normalize_date(tarih)
        tutar_k = para_kurus(tutar)
        orijinal_k = para_kurus(tutar if orijinal_tutar is None else orijinal_tutar)
        uid = self.aktif_kullanici_id
        if etiketler is None:
            self.cursor.execute(
                """
            UPDATE islemler
            SET tarih=?, tur=?, kategori=?, aciklama=?, tutar=?,
                para_birimi=?, orijinal_tutar=?
            WHERE id=? AND kullanici_id=?
            """,
                (tarih_iso, tur, kategori, aciklama, tutar_k,
                 para_birimi, orijinal_k, id, uid),
            )
        else:
            self.cursor.execute(
                """
            UPDATE islemler
            SET tarih=?, tur=?, kategori=?, aciklama=?, tutar=?, etiketler=?,
                para_birimi=?, orijinal_tutar=?
            WHERE id=? AND kullanici_id=?
            """,
                (tarih_iso, tur, kategori, aciklama, tutar_k, etiketler,
                 para_birimi, orijinal_k, id, uid),
            )
        self._log_islem("guncelle", id, f"{kategori}: {tutar}")
        self.conn.commit()

    def islemler_aralik(self, baslangic: str, bitis: str) -> List[Tuple[Any, ...]]:
        bas_iso = normalize_date(baslangic)
        bit_iso = normalize_date(bitis)
        self.cursor.execute(
            f"""
        SELECT {_ISLEM_SECIM}
        FROM islemler
        WHERE kullanici_id=? AND tarih BETWEEN ? AND ?
        ORDER BY id DESC
        """,
            (self.aktif_kullanici_id, bas_iso, bit_iso),
        )
        return self.cursor.fetchall()

    def toplam_gelir_aralik(self, baslangic: str, bitis: str) -> float:
        bas_iso = normalize_date(baslangic)
        bit_iso = normalize_date(bitis)
        self.cursor.execute(
            """
        SELECT IFNULL(SUM(tutar),0)
        FROM islemler
        WHERE kullanici_id=? AND tur='Gelir' AND tarih BETWEEN ? AND ?
        """,
            (self.aktif_kullanici_id, bas_iso, bit_iso),
        )
        row = self.cursor.fetchone()
        val = row[0] if row and row[0] is not None else 0.0
        return para_lira(val)

    def toplam_gider_aralik(self, baslangic: str, bitis: str) -> float:
        bas_iso = normalize_date(baslangic)
        bit_iso = normalize_date(bitis)
        self.cursor.execute(
            """
        SELECT IFNULL(SUM(tutar),0)
        FROM islemler
        WHERE kullanici_id=? AND tur='Gider' AND tarih BETWEEN ? AND ?
        """,
            (self.aktif_kullanici_id, bas_iso, bit_iso),
        )
        row = self.cursor.fetchone()
        val = row[0] if row and row[0] is not None else 0.0
        return para_lira(val)

    def kategori_toplamlari(self, tur: Optional[str] = None) -> List[Tuple[str, float]]:
        sorgu = (
            "SELECT kategori, SUM(tutar)/100.0 FROM islemler WHERE kullanici_id=?"
        )
        kosullar: List[Any] = [self.aktif_kullanici_id]
        if tur:
            sorgu += " AND tur=?"
            kosullar.append(tur)
        sorgu += " GROUP BY kategori ORDER BY SUM(tutar) DESC"
        self.cursor.execute(sorgu, tuple(kosullar))
        return self.cursor.fetchall()

    def aylik_ozet(self) -> List[Tuple[str, float, float]]:
        """(ay, gelir_toplam, gider_toplam) listesi döner. Son 12 ay."""
        self.cursor.execute(
            """
        SELECT
            strftime('%Y-%m', tarih) AS ay,
            SUM(CASE WHEN tur='Gelir' THEN tutar ELSE 0 END),
            SUM(CASE WHEN tur='Gider' THEN tutar ELSE 0 END)
        FROM islemler
        WHERE kullanici_id=?
        GROUP BY ay
        ORDER BY ay DESC
        LIMIT 12
        """,
            (self.aktif_kullanici_id,),
        )
        return [
            (r[0], para_lira(r[1]), para_lira(r[2]))
            for r in self.cursor.fetchall()
        ]

    def export_csv(self, path: str) -> None:
        with open(path, "w", newline="", encoding="utf-8") as dosya:
            writer = csv.writer(dosya)
            writer.writerow([
                "id", "tarih", "tur", "kategori", "aciklama", "tutar",
                "para_birimi", "orijinal_tutar",
            ])
            for satir in self.tum_islemler():
                # Tarihi GG.AA.YYYY formatına çevir
                try:
                    dt = datetime.strptime(satir[1], "%Y-%m-%d")
                    tarih_goster = dt.strftime("%d.%m.%Y")
                except ValueError:
                    tarih_goster = satir[1]
                # satir[5]=TL tutar, satir[8]=para_birimi, satir[9]=orijinal
                para_birimi = satir[8] if len(satir) > 8 else "TRY"
                orijinal = satir[9] if len(satir) > 9 else satir[5]
                writer.writerow(
                    [satir[0], tarih_goster,
                     csv_guvenli(satir[2]), csv_guvenli(satir[3]),
                     csv_guvenli(satir[4]), satir[5], para_birimi, orijinal]
                )

    @staticmethod
    def csv_satirlarini_oku(path: str) -> List[Dict[str, str]]:
        """CSV dosyasını satır sözlüklerine ayrıştırır — DB'ye DOKUNMAZ.

        Ayrı bir metot: dosya okuma/ayrıştırma (yavaş kısım) UI worker
        thread'inde yapılabilsin, DB yazımı ana thread'de kalsın. Böylece
        büyük dosya içe aktarımı arayüzü dondurmaz (bkz. satirlari_ice_aktar).
        """
        with open(path, "r", encoding="utf-8-sig") as dosya:
            return list(csv.DictReader(dosya))

    @staticmethod
    def excel_satirlarini_oku(path: str) -> List[Dict[str, str]]:
        """Excel (.xlsx) dosyasını normalize edilmiş satır sözlüklerine
        ayrıştırır — DB'ye DOKUNMAZ (worker thread'de güvenli)."""
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            satirlar_iter = ws.iter_rows(values_only=True)
            try:
                baslik = [
                    str(h).strip().lower() if h else "" for h in next(satirlar_iter)
                ]
            except StopIteration:
                return []

            # Türkçe/İngilizce başlık eşlemesi
            eslesme = {
                "tarih": "tarih", "date": "tarih",
                "tür": "tur", "tur": "tur", "type": "tur",
                "kategori": "kategori", "category": "kategori",
                "açıklama": "aciklama", "aciklama": "aciklama",
                "description": "aciklama",
                "tutar": "tutar", "amount": "tutar",
                "etiket": "etiketler", "etiketler": "etiketler", "tags": "etiketler",
            }
            indeksler = {}
            for i, h in enumerate(baslik):
                if h in eslesme:
                    indeksler[eslesme[h]] = i

            sonuc: List[Dict[str, str]] = []
            for row in satirlar_iter:
                if row is None or all(v is None for v in row):
                    continue
                satir = {
                    k: (row[indeksler[k]] if k in indeksler else "")
                    for k in ("tarih", "tur", "kategori", "aciklama", "tutar",
                              "etiketler")
                }
                sonuc.append(
                    {k: ("" if v is None else str(v)) for k, v in satir.items()}
                )
            return sonuc
        finally:
            wb.close()

    def satirlari_ice_aktar(self, satirlar: List[Dict[str, str]]) -> int:
        """Ayrıştırılmış satırları TEK transaction'da ekler (ana thread).

        Atlanan satır sayısı self.son_ice_aktarim_atlanan'da tutulur.
        """
        eklenen = 0
        atlanan = 0
        with self._transaction():
            for satir in satirlar:
                try:
                    n = self._satir_ekle_guvenli(satir)
                    eklenen += n
                    if n == 0:
                        atlanan += 1
                except (ValueError, KeyError):
                    atlanan += 1
                    continue
        self.son_ice_aktarim_atlanan = atlanan
        return eklenen

    def import_csv(self, path: str) -> int:
        """CSV dosyasından işlemleri içe aktarır (ayrıştır + ekle). Kolaylık
        sarmalayıcısı; UI donmasını önlemek için ayrıştırma ve ekleme ayrı
        çağrılabilir (bkz. csv_satirlarini_oku / satirlari_ice_aktar)."""
        return self.satirlari_ice_aktar(self.csv_satirlarini_oku(path))

    def import_excel(self, path: str) -> int:
        """Excel dosyasından işlemleri içe aktarır (ayrıştır + ekle)."""
        return self.satirlari_ice_aktar(self.excel_satirlarini_oku(path))

    @staticmethod
    def _tutar_parse(ham: Any) -> float:
        """İçe aktarımda tutarı hem sade hem Türk (1.234,56) formatından okur.

        Ayrıştırma tek kaynaktan (ui.money.para_parse) yapılır. Buradaki eski
        kopya "tek nokta" dalını hiç ele almadığı için "45.000" gibi Türk
        binlik yazımını float("45.000")=45.0 olarak, yani 1000 kat küçük
        okuyordu. İki ayrıştırıcının ayrışması sessiz para bozulması üretir.
        """
        if ham is None:
            return 0.0
        s = str(ham).strip()
        if not s or s.replace("₺", "").replace(" ", "") == "":
            return 0.0
        return para_parse(s)

    def _satir_ekle_guvenli(self, satir: Dict[str, str]) -> int:
        """CSV/Excel içe aktarımı için ortak satır doğrulama ve ekleme mantığı (commit çağırmaz)."""
        tarih = normalize_date(satir.get("tarih", ""))
        tur = satir.get("tur", "").strip()
        kategori = satir.get("kategori", "").strip()
        aciklama = satir.get("aciklama", "").strip() or None
        tutar = para_kurus(self._tutar_parse(satir.get("tutar", "0")))
        etiketler = satir.get("etiketler", "").strip()
        if tur not in ("Gelir", "Gider") or not kategori:
            return 0
        self.cursor.execute(
            "INSERT INTO islemler (tarih, tur, kategori, aciklama, tutar, "
            "etiketler, kullanici_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (tarih, tur, kategori, aciklama, tutar, etiketler,
             self.aktif_kullanici_id),
        )
        return 1

    def kaydet_butce(self, ay: int, yil: int, kategori: str, tutar: float) -> None:
        self.cursor.execute(
            """
        INSERT INTO butceler (ay, yil, kategori, tutar, kullanici_id)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(ay, yil, kategori, kullanici_id)
            DO UPDATE SET tutar=excluded.tutar
        """,
            (ay, yil, kategori, para_kurus(tutar), self.aktif_kullanici_id),
        )
        self.conn.commit()

    def butce_listele(self, ay: int, yil: int) -> List[Tuple[str, float]]:
        self.cursor.execute(
            """
        SELECT kategori, tutar/100.0 FROM butceler
        WHERE ay=? AND yil=? AND kullanici_id=?
        ORDER BY kategori
        """,
            (ay, yil, self.aktif_kullanici_id),
        )
        return self.cursor.fetchall()

    def butce_sil(self, ay: int, yil: int, kategori: str) -> None:
        self.cursor.execute(
            "DELETE FROM butceler WHERE ay=? AND yil=? AND kategori=? "
            "AND kullanici_id=?",
            (ay, yil, kategori, self.aktif_kullanici_id),
        )
        self.conn.commit()

    def butce_kopyala(
        self, kaynak_ay: int, kaynak_yil: int, hedef_ay: int, hedef_yil: int
    ) -> int:
        """Bir ayın bütçelerini başka aya kopyalar (ay devri).

        Yeni ay başladığında bütçeler 'hiç tanımlanmamış' gibi kaybolmasın
        diye önceki aydan kopyalama sağlar. Kopyalanan kalem sayısını döner.
        """
        kaynak = self.butce_listele(kaynak_ay, kaynak_yil)
        for kategori, tutar in kaynak:
            self.kaydet_butce(hedef_ay, hedef_yil, kategori, tutar)
        return len(kaynak)

    def butce_durumu(self, ay: int, yil: int) -> List[Dict[str, Any]]:
        self.cursor.execute(
            """
        SELECT
            b.kategori,
            b.tutar AS butce,
            COALESCE(
                SUM(CASE WHEN i.tur='Gider' THEN i.tutar ELSE 0 END),
                0
            ) AS harcanan
        FROM butceler b
        LEFT JOIN islemler i ON i.kategori = b.kategori
        AND i.kullanici_id = b.kullanici_id
        AND strftime('%m', i.tarih) = printf('%02d', b.ay)
        AND strftime('%Y', i.tarih) = b.yil
        WHERE b.ay=? AND b.yil=? AND b.kullanici_id=?
        GROUP BY b.kategori, b.tutar
        ORDER BY b.kategori
        """,
            (ay, yil, self.aktif_kullanici_id),
        )
        sonuc = []
        for kategori, butce, harcanan in self.cursor.fetchall():
            sonuc.append(
                {
                    "kategori": kategori,
                    "butce": para_lira(butce),
                    "harcanan": para_lira(harcanan),
                    "kalan": para_lira(int(butce) - int(harcanan)),
                }
            )
        return sonuc

    def ayar_kaydet(self, anahtar: str, deger: str) -> None:
        self.cursor.execute(
            """
        INSERT INTO ayarlar (anahtar, deger)
        VALUES (?, ?)
        ON CONFLICT(anahtar) DO UPDATE SET deger=excluded.deger
        """,
            (anahtar, deger),
        )
        self.conn.commit()

    def yedekle(self, hedef_yol: str) -> None:
        # WAL modunda veriler önce -wal dosyasına yazılır; checkpoint
        # yapılmadan ana .db dosyası kopyalanırsa yedek eksik/boş çıkar.
        # Ayrı/geçici bir cursor kullanılır: self.cursor üzerinden checkpoint
        # çağırmak Windows'ta ana .db dosyasını sonradan kilitli bırakıyor.
        self.conn.commit()
        gecici_cursor = self.conn.cursor()
        gecici_cursor.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        gecici_cursor.close()
        shutil.copy2(DB_PATH, hedef_yol)
        # Yedeğin yanına HMAC imzası yaz — düz sha256 imzasızdı, saldırgan
        # yedeği değiştirip checksum'ı yeniden üretebiliyordu. HMAC anahtarı
        # yedeklerle birlikte taşınmadığı için kurcalama tespit edilir.
        mac = hmac.new(_hmac_anahtari(), digestmod=hashlib.sha256)
        with open(hedef_yol, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                mac.update(chunk)
        with open(str(hedef_yol) + ".hmac", "w", encoding="utf-8") as cf:
            cf.write(mac.hexdigest())

    def geri_yukle(self, kaynak_yol: str) -> None:
        # Bütünlük kontrolü: önce HMAC (yeni), yoksa eski düz sha256 (geriye
        # uyumluluk). HMAC eşleşmezse kurcalanmış/bozuk kabul edilir.
        hmac_path = str(kaynak_yol) + ".hmac"
        sha_path = str(kaynak_yol) + ".sha256"
        if Path(hmac_path).exists():
            mac = hmac.new(_hmac_anahtari(), digestmod=hashlib.sha256)
            with open(kaynak_yol, "rb") as f:
                for chunk in iter(lambda: f.read(8192), b""):
                    mac.update(chunk)
            with open(hmac_path, "r", encoding="utf-8") as cf:
                expected = cf.read().strip()
            if not hmac.compare_digest(mac.hexdigest(), expected):
                raise ValueError("Yedek bütünlük (HMAC) kontrolü başarısız.")
        elif Path(sha_path).exists():
            h = hashlib.sha256()
            with open(kaynak_yol, "rb") as f:
                for chunk in iter(lambda: f.read(8192), b""):
                    h.update(chunk)
            with open(sha_path, "r", encoding="utf-8") as cf:
                expected = cf.read().strip()
            if not hmac.compare_digest(h.hexdigest(), expected):
                raise ValueError("Yedek bütünlük kontrolü başarısız.")

        # Dosyanın gerçekten bir SQLite veritabanı olduğunu doğrula —
        # bozuk/yabancı bir dosya tüm veriyi (kullanıcı tablosu dahil) yok eder.
        with open(kaynak_yol, "rb") as f:
            if f.read(16) != b"SQLite format 3\x00":
                raise ValueError("Seçilen dosya geçerli bir yedek değil.")
        gecici = sqlite3.connect(kaynak_yol)
        try:
            if gecici.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
                raise ValueError("Yedek dosyası bozuk (integrity_check).")
        finally:
            gecici.close()

        # Mevcut veritabanını üzerine yazmadan önce güvenlik yedeği al
        try:
            self.conn.commit()
            self.conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        except Exception:
            pass
        self.conn.close()

        # Eski WAL/SHM dosyaları geri yüklenen DB'nin üzerine 'replay'
        # edilip sessiz bozulmaya yol açabilir — kopyalamadan önce silinir.
        for ek in ("-wal", "-shm"):
            yan = Path(str(DB_PATH) + ek)
            if yan.exists():
                try:
                    yan.unlink()
                except OSError:
                    pass
        if Path(DB_PATH).exists():
            try:
                shutil.copy2(DB_PATH, str(DB_PATH) + ".restore-bak")
            except OSError:
                pass

        shutil.copy2(kaynak_yol, DB_PATH)
        self.conn = self._baglan()
        self.cursor = self.conn.cursor()
        # Geri yüklenen DB'nin id'leri farklı; eski geri-al geçmişi geçersiz
        self._silinen_yigin = []

        # Başarılı geri yüklemeden sonra güvenlik kopyasını SİL. Kopya tüm
        # finans geçmişini ve parola hash'lerini şifresiz, imzasız biçimde
        # süresiz bırakıyordu ve hiç temizlenmiyordu.
        guvenlik_kopyasi = Path(str(DB_PATH) + ".restore-bak")
        if guvenlik_kopyasi.exists():
            try:
                guvenlik_kopyasi.unlink()
            except OSError:
                logging.getLogger(__name__).warning(
                    "Geri yükleme güvenlik kopyası silinemedi: %s", guvenlik_kopyasi
                )

    def ayar_oku(self, anahtar: str, varsayilan: Optional[str] = None) -> Optional[str]:
        self.cursor.execute("SELECT deger FROM ayarlar WHERE anahtar=?", (anahtar,))
        sonuc = self.cursor.fetchone()
        if not sonuc:
            return varsayilan
        value: Any = sonuc[0]
        if value is None:
            return varsayilan
        return str(value)

    # ==========================
    # TOPLAM GELİR
    # ==========================

    def toplam_gelir(self) -> float:
        self.cursor.execute(
            "SELECT IFNULL(SUM(tutar),0) FROM islemler "
            "WHERE tur='Gelir' AND kullanici_id=?",
            (self.aktif_kullanici_id,),
        )
        row = self.cursor.fetchone()
        val = row[0] if row and row[0] is not None else 0.0
        return para_lira(val)

    # ==========================
    # TOPLAM GİDER
    # ==========================

    def toplam_gider(self) -> float:
        self.cursor.execute(
            "SELECT IFNULL(SUM(tutar),0) FROM islemler "
            "WHERE tur='Gider' AND kullanici_id=?",
            (self.aktif_kullanici_id,),
        )
        row = self.cursor.fetchone()
        val = row[0] if row and row[0] is not None else 0.0
        return para_lira(val)

    # ==========================
    # BAKİYE
    # ==========================

    def bakiye(self) -> float:
        return self.toplam_gelir() - self.toplam_gider()

    # ==========================
    # İŞLEM SİL
    # ==========================

    def sil(self, islem_id: int) -> None:
        """Tek işlem siler ve geri-al yığınına tek kalemlik parti olarak iter."""
        self.sil_toplu([islem_id])

    def sil_toplu(self, islem_idler: List[int]) -> int:
        """Birden çok işlemi TEK transaction'da siler; TEK geri-al birimi olur.

        Önceden UI döngüde sil() çağırıyordu ve _son_silinen tek slot olduğu
        için 8 kayıt silinince yalnızca sonuncusu geri alınabiliyordu —
        kalan 7'si mesaj tekil olduğundan fark edilmeden kalıcı kayboluyordu.
        Silinen satır sayısını döner.
        """
        uid = self.aktif_kullanici_id
        parti: List[Any] = []
        with self._transaction():
            for islem_id in islem_idler:
                self.cursor.execute(
                    "SELECT * FROM islemler WHERE id=? AND kullanici_id=?",
                    (islem_id, uid),
                )
                row = self.cursor.fetchone()
                if row is None:
                    continue
                parti.append(row)
                self.cursor.execute(
                    "DELETE FROM islemler WHERE id=? AND kullanici_id=?",
                    (islem_id, uid),
                )
                self._log_islem("sil", islem_id, "İşlem silindi")
        if parti:
            self._silinen_yigin.append(parti)
            # Yığını sınırla: geri-al geçmişi belleği süresiz büyütmemeli
            del self._silinen_yigin[:-GERI_AL_YIGIN_SINIRI]
        return len(parti)

    def geri_al(self) -> int:
        """Son silme partisini geri getirir; geri gelen kayıt sayısını döner.

        Dönüş int: 0 (falsy) 'geri alınacak bir şey yok' demektir, böylece
        mevcut `if db.geri_al():` çağrıları çalışmaya devam eder.
        """
        if not self._silinen_yigin:
            return 0
        parti = self._silinen_yigin[-1]
        try:
            with self._transaction():
                for veri in parti:
                    # SELECT * tüm sütunları ham (kuruş) verir; hepsini
                    # koruyarak geri ekle. 10 sütun = çoklu para birimi şeması
                    # (para_birimi + orijinal_tutar); yoksa USD işlem geri
                    # alınınca TL'ye düşerdi.
                    if len(veri) >= 10:
                        self.cursor.execute(
                            "INSERT INTO islemler (id, tarih, tur, kategori, "
                            "aciklama, tutar, etiketler, kullanici_id, "
                            "para_birimi, orijinal_tutar) "
                            "VALUES (?,?,?,?,?,?,?,?,?,?)",
                            veri[:10],
                        )
                    elif len(veri) >= 8:
                        # kullanici_id'yi de koru (8. sütun), yoksa admin'e düşer
                        self.cursor.execute(
                            "INSERT INTO islemler (id, tarih, tur, kategori, "
                            "aciklama, tutar, etiketler, kullanici_id) "
                            "VALUES (?,?,?,?,?,?,?,?)",
                            veri[:8],
                        )
                    elif len(veri) >= 7:
                        self.cursor.execute(
                            "INSERT INTO islemler (id, tarih, tur, kategori, "
                            "aciklama, tutar, etiketler) VALUES (?,?,?,?,?,?,?)",
                            veri[:7],
                        )
                    else:
                        # Eski DB'lerde etiketler sütunu olmayabilir
                        self.cursor.execute(
                            "INSERT INTO islemler (id, tarih, tur, kategori, "
                            "aciklama, tutar) VALUES (?,?,?,?,?,?)",
                            veri[:6],
                        )
        except Exception:
            return 0
        self._silinen_yigin.pop()
        return len(parti)

    # ==========================
    # AYLIK KARŞILAŞTIRMA
    # ==========================

    def aylik_karsilastirma(self) -> Dict[str, Any]:
        """Bu ay ve geçen ayın gelir/gider karşılaştırması."""
        from datetime import datetime
        simdi = datetime.now()
        bu_ay = simdi.month
        bu_yil = simdi.year
        gecen_ay = 12 if bu_ay == 1 else bu_ay - 1
        gecen_yil = bu_yil - 1 if bu_ay == 1 else bu_yil

        def _ay_toplam(ay, yil, tur):
            self.cursor.execute(
                "SELECT COALESCE(SUM(tutar),0) FROM islemler "
                "WHERE kullanici_id=? AND tur=? "
                "AND CAST(strftime('%m', tarih) AS INTEGER)=? "
                "AND CAST(strftime('%Y', tarih) AS INTEGER)=?",
                (self.aktif_kullanici_id, tur, ay, yil),
            )
            row = self.cursor.fetchone()
            return para_lira(row[0]) if row else 0.0

        return {
            "bu_ay": {"ay": bu_ay, "yil": bu_yil,
                      "gelir": _ay_toplam(bu_ay, bu_yil, "Gelir"),
                      "gider": _ay_toplam(bu_ay, bu_yil, "Gider")},
            "gecen_ay": {"ay": gecen_ay, "yil": gecen_yil,
                         "gelir": _ay_toplam(gecen_ay, gecen_yil, "Gelir"),
                         "gider": _ay_toplam(gecen_ay, gecen_yil, "Gider")},
        }

    def yillik_karsilastirma(self) -> List[Tuple[str, float, float]]:
        """(yil, gelir_toplam, gider_toplam) listesi döner — tüm yıllar, eskiden yeniye."""
        self.cursor.execute(
            """
        SELECT
            strftime('%Y', tarih) AS yil,
            COALESCE(SUM(CASE WHEN tur='Gelir' THEN tutar ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN tur='Gider' THEN tutar ELSE 0 END), 0)
        FROM islemler
        WHERE kullanici_id=?
        GROUP BY yil
        ORDER BY yil ASC
        """,
            (self.aktif_kullanici_id,),
        )
        return [
            (r[0], para_lira(r[1]), para_lira(r[2]))
            for r in self.cursor.fetchall()
        ]

    # ==========================
    # GÜNLÜK / HAFTALIK FİLTRE
    # ==========================

    def gunluk_islemler(self) -> List[Tuple[Any, ...]]:
        from datetime import date
        bugun = date.today().strftime("%Y-%m-%d")
        self.cursor.execute(
            f"SELECT {_ISLEM_SECIM} FROM islemler WHERE tarih=? "
            "AND kullanici_id=? ORDER BY id DESC",
            (bugun, self.aktif_kullanici_id),
        )
        return self.cursor.fetchall()

    def haftalik_islemler(self) -> List[Tuple[Any, ...]]:
        from datetime import date, timedelta
        bugun = date.today()
        hafta_basi = (bugun - timedelta(days=bugun.weekday())).strftime("%Y-%m-%d")
        bugun_str = bugun.strftime("%Y-%m-%d")
        self.cursor.execute(
            f"SELECT {_ISLEM_SECIM} FROM islemler WHERE tarih BETWEEN ? AND ? "
            "AND kullanici_id=? ORDER BY id DESC",
            (hafta_basi, bugun_str, self.aktif_kullanici_id),
        )
        return self.cursor.fetchall()

    # ==========================
    # KAPAT
    # ==========================

    def close(self) -> None:
        self.conn.close()
